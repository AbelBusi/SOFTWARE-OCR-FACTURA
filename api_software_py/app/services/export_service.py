import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)


COLUMNAS_GENERAL = [
    "N° Comprobante",
    "Fecha",
    "Empresa",
    "RUC",
    "Subtotal",
    "IGV",
    "Total",
    "Estado",
]

COLUMNAS_DETALLE = [
    "Descripción",
    "Cantidad",
    "Precio Unitario",
    "Subtotal",
]

MONETARIAS_GENERAL = {"Subtotal", "IGV", "Total"}

_COLOR_CABECERA = colors.HexColor("#263238")
_COLOR_FILA = colors.HexColor("#F2F5F9")
_COLOR_BORDE = colors.HexColor("#CFD8DC")
_COLOR_TEXTO = colors.HexColor("#263238")
_COLOR_TOTAL = colors.HexColor("#E3F2FD")

MIME = {
    "pdf": ("application/pdf", "pdf"),
    "excel": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    ),
}


class ExportService:

    def __init__(self):
        self._general = {
            "pdf": self._general_pdf,
            "excel": self._general_excel,
        }
        self._individual = {
            "pdf": self._individual_pdf,
            "excel": self._individual_excel,
        }

    def formatos_disponibles(self):
        return list(self._general.keys())

    def media_type(self, formato):
        return MIME[formato][0]

    def extension(self, formato):
        return MIME[formato][1]

    def exportar_general(self, filas, formato, q=None, fecha=None) -> bytes:
        generador = self._general.get(formato)
        if generador is None:
            raise ValueError(f"Formato no soportado: {formato}")
        return generador(filas, q, fecha)

    def exportar_individual(self, factura, formato) -> bytes:
        generador = self._individual.get(formato)
        if generador is None:
            raise ValueError(f"Formato no soportado: {formato}")
        return generador(factura)

    def _num(self, dato):
        return float(dato) if dato is not None else 0.0

    def _money(self, valor):
        return f"S/ {self._num(valor):,.2f}"

    def _proveedor(self, empresa):
        if empresa and empresa.razon_social:
            return empresa.razon_social
        return "-"

    def _ruc(self, empresa):
        if empresa and empresa.ruc and not empresa.ruc.startswith("SINRUC-"):
            return empresa.ruc
        return "-"

    def _direccion(self, empresa):
        if empresa and empresa.direccion:
            return empresa.direccion
        return "-"

    def _fecha(self, factura):
        return factura.fecha_emision.strftime("%Y-%m-%d") if factura.fecha_emision else "-"

    def _estado(self, factura):
        return "Activo" if getattr(factura, "estado", 1) == 1 else "Eliminado"

    def _observaciones(self, factura):
        return getattr(factura, "observaciones", None)

    def _filtros_texto(self, q, fecha):
        partes = []
        if q:
            partes.append(f'Búsqueda: "{q}"')
        if fecha:
            partes.append(f"Fecha: {fecha}")
        return " | ".join(partes) if partes else "Sin filtros aplicados"

    def _estilos(self):
        return {
            "titulo": ParagraphStyle(
                "titulo", fontName="Helvetica-Bold", fontSize=16,
                textColor=_COLOR_TEXTO, spaceAfter=2,
            ),
            "meta": ParagraphStyle(
                "meta", fontName="Helvetica", fontSize=9,
                textColor=colors.HexColor("#607D8B"), leading=13,
            ),
            "celda": ParagraphStyle(
                "celda", fontName="Helvetica", fontSize=8,
                textColor=_COLOR_TEXTO, leading=10,
            ),
            "celda_der": ParagraphStyle(
                "celda_der", fontName="Helvetica", fontSize=8,
                textColor=_COLOR_TEXTO, leading=10, alignment=TA_RIGHT,
            ),
            "cab": ParagraphStyle(
                "cab", fontName="Helvetica-Bold", fontSize=8.5,
                textColor=colors.white, alignment=TA_LEFT,
            ),
            "cab_der": ParagraphStyle(
                "cab_der", fontName="Helvetica-Bold", fontSize=8.5,
                textColor=colors.white, alignment=TA_RIGHT,
            ),
        }

    def _pie(self, canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#90A4AE"))
        canvas.drawString(15 * mm, 8 * mm, "OCR Factura - Reporte de Comprobantes")
        canvas.drawRightString(
            doc.pagesize[0] - 15 * mm, 8 * mm, f"Página {canvas.getPageNumber()}"
        )
        canvas.restoreState()

    def _general_pdf(self, filas, q, fecha) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            leftMargin=15 * mm, rightMargin=15 * mm,
            topMargin=15 * mm, bottomMargin=15 * mm,
            title="Reporte de Comprobantes",
        )
        s = self._estilos()
        generado = datetime.now().strftime("%Y-%m-%d %H:%M")

        story = [
            Paragraph("Reporte de Comprobantes", s["titulo"]),
            Paragraph(f"Generado: {generado}", s["meta"]),
            Paragraph(self._filtros_texto(q, fecha), s["meta"]),
            Paragraph(f"Total de comprobantes: {len(filas)}", s["meta"]),
            Spacer(1, 8 * mm),
        ]

        cabecera = [
            Paragraph(c, s["cab_der"] if c in MONETARIAS_GENERAL else s["cab"])
            for c in COLUMNAS_GENERAL
        ]
        tabla = [cabecera]

        total_general = 0.0
        for factura, empresa in filas:
            total_general += self._num(factura.total)
            tabla.append([
                Paragraph(factura.numero_comprobante or "-", s["celda"]),
                Paragraph(self._fecha(factura), s["celda"]),
                Paragraph(self._proveedor(empresa), s["celda"]),
                Paragraph(self._ruc(empresa), s["celda"]),
                Paragraph(self._money(factura.subtotal), s["celda_der"]),
                Paragraph(self._money(factura.igv), s["celda_der"]),
                Paragraph(self._money(factura.total), s["celda_der"]),
                Paragraph(self._estado(factura), s["celda"]),
            ])

        tabla.append([
            Paragraph("TOTAL", ParagraphStyle("t", parent=s["celda"], fontName="Helvetica-Bold")),
            "", "", "", "", "",
            Paragraph(self._money(total_general),
                      ParagraphStyle("td", parent=s["celda_der"], fontName="Helvetica-Bold")),
            "",
        ])

        anchos = [34 * mm, 24 * mm, 62 * mm, 28 * mm, 28 * mm, 26 * mm, 30 * mm, 30 * mm]
        t = Table(tabla, colWidths=anchos, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _COLOR_CABECERA),
            ("TOPPADDING", (0, 0), (-1, 0), 6),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, _COLOR_FILA]),
            ("GRID", (0, 0), (-1, -1), 0.4, _COLOR_BORDE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("SPAN", (0, -1), (5, -1)),
            ("SPAN", (6, -1), (7, -1)),
            ("BACKGROUND", (0, -1), (-1, -1), _COLOR_TOTAL),
            ("LINEABOVE", (0, -1), (-1, -1), 0.8, _COLOR_CABECERA),
        ]))
        story.append(t)
        doc.build(story, onFirstPage=self._pie, onLaterPages=self._pie)
        return buffer.getvalue()

    def _individual_pdf(self, factura) -> bytes:
        empresa = factura.empresa
        detalles = factura.detalles
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=18 * mm, rightMargin=18 * mm,
            topMargin=18 * mm, bottomMargin=18 * mm,
            title=f"Comprobante {factura.numero_comprobante or ''}".strip(),
        )
        s = self._estilos()
        generado = datetime.now().strftime("%Y-%m-%d %H:%M")

        etiqueta = ParagraphStyle("etq", fontName="Helvetica-Bold", fontSize=9, textColor=_COLOR_TEXTO)
        dato = ParagraphStyle("dato", fontName="Helvetica", fontSize=9, textColor=_COLOR_TEXTO)

        def bloque(titulo, pares):
            filas = [[Paragraph(t, etiqueta), Paragraph(v, dato)] for t, v in pares]
            tabla = Table(filas, colWidths=[42 * mm, 130 * mm])
            tabla.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), _COLOR_FILA),
                ("GRID", (0, 0), (-1, -1), 0.4, _COLOR_BORDE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]))
            return [Paragraph(titulo, ParagraphStyle("sec", fontName="Helvetica-Bold",
                    fontSize=11, textColor=_COLOR_TEXTO, spaceAfter=4)), tabla, Spacer(1, 6 * mm)]

        story = [
            Paragraph("Detalle de Comprobante", s["titulo"]),
            Paragraph(f"Generado: {generado}", s["meta"]),
            Spacer(1, 6 * mm),
        ]

        story += bloque("Empresa", [
            ("Razón social", self._proveedor(empresa)),
            ("RUC", self._ruc(empresa)),
            ("Dirección", self._direccion(empresa)),
        ])

        datos_factura = [
            ("Tipo", factura.tipo_comprobante or "-"),
            ("N° Comprobante", factura.numero_comprobante or "-"),
            ("Fecha de emisión", self._fecha(factura)),
            ("Estado", self._estado(factura)),
        ]
        observaciones = self._observaciones(factura)
        if observaciones:
            datos_factura.append(("Observaciones", observaciones))
        story += bloque("Comprobante", datos_factura)

        story.append(Paragraph("Productos / Servicios", ParagraphStyle(
            "sec2", fontName="Helvetica-Bold", fontSize=11, textColor=_COLOR_TEXTO, spaceAfter=4)))

        cabecera = [
            Paragraph(c, s["cab_der"] if c in ("Cantidad", "Precio Unitario", "Subtotal") else s["cab"])
            for c in COLUMNAS_DETALLE
        ]
        tabla = [cabecera]
        if detalles:
            for d in detalles:
                cantidad = self._num(d.cantidad)
                cantidad_txt = str(int(cantidad)) if cantidad == int(cantidad) else f"{cantidad:g}"
                tabla.append([
                    Paragraph(d.descripcion or "-", s["celda"]),
                    Paragraph(cantidad_txt, s["celda_der"]),
                    Paragraph(self._money(d.precio_unitario), s["celda_der"]),
                    Paragraph(self._money(d.subtotal), s["celda_der"]),
                ])
        else:
            tabla.append([Paragraph("Sin ítems detallados", s["celda"]), "", "", ""])

        t = Table(tabla, colWidths=[92 * mm, 26 * mm, 30 * mm, 26 * mm], repeatRows=1)
        estilo = [
            ("BACKGROUND", (0, 0), (-1, 0), _COLOR_CABECERA),
            ("TOPPADDING", (0, 0), (-1, 0), 6),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("GRID", (0, 0), (-1, -1), 0.4, _COLOR_BORDE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]
        if detalles:
            estilo.append(("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _COLOR_FILA]))
        else:
            estilo.append(("SPAN", (0, 1), (-1, 1)))
        t.setStyle(TableStyle(estilo))
        story.append(t)
        story.append(Spacer(1, 6 * mm))

        totales = Table([
            ["Subtotal", self._money(factura.subtotal)],
            ["IGV", self._money(factura.igv)],
            ["Total", self._money(factura.total)],
        ], colWidths=[122 * mm, 52 * mm])
        totales.setStyle(TableStyle([
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("TEXTCOLOR", (0, 0), (-1, -1), _COLOR_TEXTO),
            ("LINEABOVE", (0, -1), (-1, -1), 0.8, _COLOR_CABECERA),
            ("BACKGROUND", (0, -1), (-1, -1), _COLOR_TOTAL),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(totales)

        doc.build(story, onFirstPage=self._pie, onLaterPages=self._pie)
        return buffer.getvalue()

    def _estilo_excel(self):
        return {
            "fill_cab": PatternFill("solid", fgColor="263238"),
            "font_cab": Font(bold=True, color="FFFFFF", size=10),
            "font_titulo": Font(bold=True, size=14, color="263238"),
            "font_meta": Font(size=9, color="607D8B"),
            "font_etq": Font(bold=True, color="263238"),
            "borde": Border(*[Side(style="thin", color="CFD8DC")] * 4),
            "centro": Alignment(horizontal="center", vertical="center"),
            "derecha": Alignment(horizontal="right"),
        }

    def _general_excel(self, filas, q, fecha) -> bytes:
        est = self._estilo_excel()
        wb = Workbook()
        ws = wb.active
        ws.title = "Comprobantes"
        n_col = len(COLUMNAS_GENERAL)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_col)
        ws.cell(row=1, column=1, value="Reporte de Comprobantes").font = est["font_titulo"]
        generado = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_col)
        ws.cell(row=2, column=1, value=f"Generado: {generado}").font = est["font_meta"]
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_col)
        ws.cell(row=3, column=1, value=self._filtros_texto(q, fecha)).font = est["font_meta"]

        fila_cab = 5
        for col, nombre in enumerate(COLUMNAS_GENERAL, start=1):
            celda = ws.cell(row=fila_cab, column=col, value=nombre)
            celda.font = est["font_cab"]
            celda.fill = est["fill_cab"]
            celda.alignment = est["centro"]
            celda.border = est["borde"]

        total_general = 0.0
        fila = fila_cab + 1
        for factura, empresa in filas:
            total_general += self._num(factura.total)
            valores = [
                factura.numero_comprobante or "-",
                self._fecha(factura),
                self._proveedor(empresa),
                self._ruc(empresa),
                self._num(factura.subtotal),
                self._num(factura.igv),
                self._num(factura.total),
                self._estado(factura),
            ]
            for col, valor in enumerate(valores, start=1):
                celda = ws.cell(row=fila, column=col, value=valor)
                celda.border = est["borde"]
                if COLUMNAS_GENERAL[col - 1] in MONETARIAS_GENERAL:
                    celda.number_format = '"S/ "#,##0.00'
                    celda.alignment = est["derecha"]
            fila += 1

        ws.cell(row=fila, column=1, value="TOTAL").font = Font(bold=True)
        celda_total = ws.cell(row=fila, column=7, value=total_general)
        celda_total.font = Font(bold=True)
        celda_total.number_format = '"S/ "#,##0.00'
        celda_total.alignment = est["derecha"]

        anchos = [20, 14, 40, 16, 14, 12, 14, 16]
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho
        ws.freeze_panes = f"A{fila_cab + 1}"

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _individual_excel(self, factura) -> bytes:
        empresa = factura.empresa
        detalles = factura.detalles
        est = self._estilo_excel()
        wb = Workbook()
        ws = wb.active
        ws.title = "Comprobante"

        ws.merge_cells("A1:D1")
        ws.cell(row=1, column=1, value="Detalle de Comprobante").font = est["font_titulo"]
        generado = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.merge_cells("A2:D2")
        ws.cell(row=2, column=1, value=f"Generado: {generado}").font = est["font_meta"]

        fila = 4

        def escribir_par(etiqueta, valor):
            nonlocal fila
            c1 = ws.cell(row=fila, column=1, value=etiqueta)
            c1.font = est["font_etq"]
            c1.fill = est["fill_cab"]
            c1.font = Font(bold=True, color="FFFFFF")
            ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=4)
            ws.cell(row=fila, column=2, value=valor)
            fila += 1

        ws.cell(row=fila, column=1, value="EMPRESA").font = Font(bold=True, size=11, color="263238")
        fila += 1
        escribir_par("Razón social", self._proveedor(empresa))
        escribir_par("RUC", self._ruc(empresa))
        escribir_par("Dirección", self._direccion(empresa))
        fila += 1

        ws.cell(row=fila, column=1, value="COMPROBANTE").font = Font(bold=True, size=11, color="263238")
        fila += 1
        escribir_par("Tipo", factura.tipo_comprobante or "-")
        escribir_par("N° Comprobante", factura.numero_comprobante or "-")
        escribir_par("Fecha de emisión", self._fecha(factura))
        escribir_par("Estado", self._estado(factura))
        observaciones = self._observaciones(factura)
        if observaciones:
            escribir_par("Observaciones", observaciones)
        fila += 1

        ws.cell(row=fila, column=1, value="PRODUCTOS / SERVICIOS").font = Font(bold=True, size=11, color="263238")
        fila += 1

        for col, nombre in enumerate(COLUMNAS_DETALLE, start=1):
            celda = ws.cell(row=fila, column=col, value=nombre)
            celda.font = est["font_cab"]
            celda.fill = est["fill_cab"]
            celda.alignment = est["centro"]
            celda.border = est["borde"]
        fila += 1

        if detalles:
            for d in detalles:
                valores = [
                    d.descripcion or "-",
                    self._num(d.cantidad),
                    self._num(d.precio_unitario),
                    self._num(d.subtotal),
                ]
                for col, valor in enumerate(valores, start=1):
                    celda = ws.cell(row=fila, column=col, value=valor)
                    celda.border = est["borde"]
                    if col >= 3:
                        celda.number_format = '"S/ "#,##0.00'
                        celda.alignment = est["derecha"]
                fila += 1
        else:
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
            ws.cell(row=fila, column=1, value="Sin ítems detallados")
            fila += 1

        fila += 1
        for etiqueta, valor in (
            ("Subtotal", self._num(factura.subtotal)),
            ("IGV", self._num(factura.igv)),
            ("Total", self._num(factura.total)),
        ):
            ws.cell(row=fila, column=3, value=etiqueta).font = Font(bold=(etiqueta == "Total"))
            celda = ws.cell(row=fila, column=4, value=valor)
            celda.number_format = '"S/ "#,##0.00'
            celda.alignment = est["derecha"]
            celda.font = Font(bold=(etiqueta == "Total"))
            fila += 1

        anchos = [40, 16, 18, 16]
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
